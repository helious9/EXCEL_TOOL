import openpyxl
from googletrans import Translator, LANGUAGES
import time
from typing import Dict, Tuple
import re


def translate_excel_sheets(
        input_file: str,
        output_file: str,
        src_lang: str = 'zh-cn',
        tgt_lang: str = 'en',
        delay: float = 0.1  # 延迟时间，避免请求过快
) -> None:
    """
    翻译Excel文件中所有sheet的中文内容

    参数:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径
        src_lang: 源语言 (默认: 简体中文)
        tgt_lang: 目标语言 (默认: 英语)
        delay: 请求延迟时间，避免触发API限制
    """

    # 初始化翻译器
    translator = Translator()

    # 加载Excel文件
    print(f"正在加载文件: {input_file}")
    workbook = openpyxl.load_workbook(input_file)

    # 用于缓存翻译结果，避免重复翻译相同内容
    translation_cache: Dict[str, str] = {}

    # 翻译文本的函数
    def translate_text(text: str) -> str:
        """翻译单个文本，使用缓存"""
        if not text or not isinstance(text, str):
            return text

        # 清理文本中的多余空格和换行
        text = text.strip()

        # 检查缓存
        if text in translation_cache:
            return translation_cache[text]

        try:
            # 翻译文本
            translated = translator.translate(text, src=src_lang, dest=tgt_lang)
            result = f"{translated.text}\n{text}"

            # 存入缓存
            translation_cache[text] = result

            # 添加延迟避免请求过快
            time.sleep(delay)

            return result

        except Exception as e:
            print(f"翻译失败: '{text}' - 错误: {e}")
            return text

    # 遍历所有sheet
    for sheet_name in workbook.sheetnames:
        print(f"正在处理sheet: {sheet_name}")
        sheet = workbook[sheet_name]

        # 获取所有合并单元格的范围
        merged_ranges = list(sheet.merged_cells.ranges)

        # 创建一个集合来存储合并单元格的位置
        merged_cells = set()
        for merged_range in merged_ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells.add((row, col))

        # 记录每个合并区域的主单元格（左上角）
        merged_master_cells = {}
        for merged_range in merged_ranges:
            master_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_master_cells[(row, col)] = master_cell

        # 遍历所有有内容的单元格
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                # 跳过合并单元格中的非主单元格
                if (row, col) in merged_cells:
                    # 如果是合并单元格但不是主单元格，跳过
                    if (row, col) in merged_master_cells:
                        cell = merged_master_cells[(row, col)]
                        # 检查是否已经处理过这个合并区域
                        if row != merged_master_cells[(row, col)].row or col != merged_master_cells[(row, col)].column:
                            continue
                    else:
                        continue
                else:
                    cell = sheet.cell(row=row, column=col)

                # 检查单元格是否有内容
                if cell.value and isinstance(cell.value, str):
                    # 检查单元格内容是否包含中文
                    if re.search(r'[\u4e00-\u9fff]', cell.value):
                        try:
                            # 翻译内容
                            translated_content = translate_text(cell.value)
                            cell.value = translated_content

                            # 设置单元格格式为自动换行
                            cell.alignment = openpyxl.styles.Alignment(
                                wrapText=True,
                                vertical='top'
                            )

                        except Exception as e:
                            print(f"处理单元格 {cell.coordinate} 时出错: {e}")
                            continue

        # 调整列宽以适应新内容
        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)

            for row in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value:
                    # 检查是否为合并单元格的非主单元格
                    if (row, col_idx) in merged_cells:
                        if (row, col_idx) in merged_master_cells:
                            master_cell = merged_master_cells[(row, col_idx)]
                            if master_cell.row != row or master_cell.column != col_idx:
                                continue

                    # 计算单元格内容的最大长度
                    if isinstance(cell.value, str):
                        cell_length = max(len(line) for line in cell.value.split('\n'))
                        if cell_length > max_length:
                            max_length = cell_length

            # 设置列宽
            if max_length > 0:
                adjusted_width = min(max_length + 2, 50)  # 限制最大列宽为50
                sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存结果到新文件
    print(f"正在保存结果到: {output_file}")
    workbook.save(output_file)
    print("翻译完成！")

    # 输出统计信息
    print(f"\n翻译统计:")
    print(f"- 共处理了 {len(workbook.sheetnames)} 个sheet")
    print(f"- 缓存了 {len(translation_cache)} 个唯一翻译")


def translate_selected_sheets(
        input_file: str,
        output_file: str,
        sheet_names: list,
        src_lang: str = 'zh-cn',
        tgt_lang: str = 'en'
) -> None:
    """
    翻译Excel文件中指定的sheet

    参数:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径
        sheet_names: 要翻译的sheet名称列表
        src_lang: 源语言
        tgt_lang: 目标语言
    """

    # 初始化翻译器
    translator = Translator()

    # 加载Excel文件
    workbook = openpyxl.load_workbook(input_file)

    translation_cache = {}

    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            print(f"正在处理sheet: {sheet_name}")
            sheet = workbook[sheet_name]

            # 处理合并单元格
            merged_ranges = list(sheet.merged_cells.ranges)
            merged_cells = set()
            for merged_range in merged_ranges:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_cells.add((row, col))

            # 只处理主单元格
            merged_master_cells = {}
            for merged_range in merged_ranges:
                master_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_master_cells[(row, col)] = master_cell

            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    # 跳过合并单元格的非主单元格
                    if (row, col) in merged_cells:
                        if (row, col) in merged_master_cells:
                            cell = merged_master_cells[(row, col)]
                            if row != cell.row or col != cell.column:
                                continue
                        else:
                            continue
                    else:
                        cell = sheet.cell(row=row, column=col)

                    if cell.value and isinstance(cell.value, str):
                        # 检查是否包含中文
                        if re.search(r'[\u4e00-\u9fff]', cell.value):
                            text = cell.value.strip()

                            if text in translation_cache:
                                cell.value = translation_cache[text]
                            else:
                                try:
                                    translated = translator.translate(text, src=src_lang, dest=tgt_lang)
                                    result = f"{translated.text}\n{text}"
                                    translation_cache[text] = result
                                    cell.value = result

                                    # 设置自动换行
                                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
                                    time.sleep(0.1)

                                except Exception as e:
                                    print(f"翻译失败: '{text}' - 错误: {e}")
                                    continue

    workbook.save(output_file)
    print(f"指定sheet翻译完成！已保存到: {output_file}")


def main():
    """主函数，提供使用示例"""
    # 使用示例
    input_excel = "案例：CG02项目测试用例 .xlsx"  # 输入文件路径
    output_excel = "CG02_translated.xlsx"  # 输出文件路径

    try:
        # 方法1: 翻译所有sheet
        translate_excel_sheets(input_excel, output_excel)

        # 方法2: 仅翻译指定的sheet
        # sheets_to_translate = ["Sheet1", "Sheet2"]
        # translate_selected_sheets(input_excel, output_excel, sheets_to_translate)

    except Exception as e:
        print(f"程序执行出错: {e}")
        print("请检查文件路径和格式是否正确")


if __name__ == "__main__":
    main()